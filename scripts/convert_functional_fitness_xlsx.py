from __future__ import annotations

import csv
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "database" / "Functional+Fitness+Exercise+Database+(version+2.9).xlsx"
OUT_CSV = ROOT / "database" / "functional_fitness.exercises.supabase.csv"
OUT_AUDIT = ROOT / "database" / "functional_fitness.exercises.audit.md"


HEADER_ROW_1_INDEXED = 16  # Determined by inspection; row containing the bold headers.
SHEET_NAME = "Exercises"


ARRAY_COLUMNS = {
    "muscle_groups",
    "equipment_needed",
    "instructions",
    "tips",
}


@dataclass
class LinkCounts:
    short_links: int = 0
    long_links: int = 0


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _clean_str(value: Any) -> str:
    if _is_empty(value):
        return ""
    return str(value).strip()


def _dedupe_case_insensitive(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        s = _clean_str(item)
        if not s:
            continue
        key = s.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    return out


def _pg_array_literal(items: Iterable[str]) -> str:
    escaped: list[str] = []
    for item in items:
        s = str(item)
        s = s.replace("\\", "\\\\").replace('"', '\\"')
        escaped.append(f'"{s}"')
    return "{" + ",".join(escaped) + "}"


def _map_difficulty(level: str) -> str:
    s = _clean_str(level).casefold()
    if not s:
        return ""
    if s in {"beginner", "novice"}:
        return "beginner"
    if s == "intermediate":
        return "intermediate"
    if s in {"advanced", "expert", "master", "grand master", "legendary"}:
        return "advanced"
    return ""


def _map_category(body_region: str, prime_mover: str, target_group: str) -> str:
    br = _clean_str(body_region).casefold()
    pm = _clean_str(prime_mover)
    tg = _clean_str(target_group)

    if br == "lower body":
        return "legs"
    if br == "core":
        return "core"
    if br == "full body":
        return "full_body"

    # Upper body: infer from muscles
    upper_hint = f"{pm} {tg}".casefold()

    if any(k in upper_hint for k in ["pectoralis", "chest"]):
        return "chest"
    if any(k in upper_hint for k in ["latissimus", "trapezi", "rhombo", "erector", "back"]):
        return "back"
    if any(k in upper_hint for k in ["deltoid", "shoulder", "rotator"]):
        return "shoulders"
    if any(k in upper_hint for k in ["biceps", "triceps", "forearm", "wrist", "brachii"]):
        return "arms"

    # Fallback
    if br == "upper body":
        return "arms"

    return "full_body"


def _stable_exercise_id(key: str) -> str:
    # Stable across runs; good for repeatable imports.
    return str(uuid.uuid5(uuid.NAMESPACE_URL, f"functional-fitness:{key}"))


def _extract_youtube_links() -> tuple[list[str], list[str], LinkCounts]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [ws.cell(row=HEADER_ROW_1_INDEXED, column=c).value for c in range(1, ws.max_column + 1)]

    def find_col(header_name: str) -> int | None:
        for idx, v in enumerate(headers, start=1):
            if v == header_name:
                return idx
        return None

    col_short = find_col("Short YouTube Demonstration")
    col_long = find_col("In-Depth YouTube Explanation")

    if not col_short and not col_long:
        return [], [], LinkCounts()

    short_urls: list[str] = []
    long_urls: list[str] = []
    counts = LinkCounts()

    # Data starts right after header row.
    for r in range(HEADER_ROW_1_INDEXED + 1, ws.max_row + 1):
        if col_short:
            cell = ws.cell(row=r, column=col_short)
            url = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else ""
            short_urls.append(url or "")
            if url:
                counts.short_links += 1
        else:
            short_urls.append("")

        if col_long:
            cell = ws.cell(row=r, column=col_long)
            url = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else ""
            long_urls.append(url or "")
            if url:
                counts.long_links += 1
        else:
            long_urls.append("")

    return short_urls, long_urls, counts


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing XLSX: {XLSX_PATH}")

    # Read the sheet with the real header row
    df = pd.read_excel(XLSX_PATH, sheet_name=SHEET_NAME, skiprows=HEADER_ROW_1_INDEXED - 1)

    # Drop the first empty/unnamed column
    if len(df.columns) > 0 and str(df.columns[0]).startswith("Unnamed"):
        df = df.drop(columns=[df.columns[0]])

    # Normalize column names (strip trailing spaces)
    df.columns = [str(c).strip() for c in df.columns]

    # Extract hyperlink URLs from the sheet (pandas reads only the display text)
    short_urls, long_urls, link_counts = _extract_youtube_links()
    # Align length: openpyxl iterates to max_row; pandas read may stop earlier on empty trailing rows
    n = len(df)
    df["__short_youtube_url"] = (short_urls[:n] if len(short_urls) >= n else short_urls + [""] * (n - len(short_urls)))
    df["__long_youtube_url"] = (long_urls[:n] if len(long_urls) >= n else long_urls + [""] * (n - len(long_urls)))

    # Clean key headers (some include trailing spaces in source)
    target_col = "Target Muscle Group"
    if "Target Muscle Group" not in df.columns and "Target Muscle Group " in df.columns:
        df = df.rename(columns={"Target Muscle Group ": "Target Muscle Group"})
    if "Primary Equipment" not in df.columns and "Primary Equipment " in df.columns:
        df = df.rename(columns={"Primary Equipment ": "Primary Equipment"})

    now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S+00")

    out_rows: list[dict[str, Any]] = []
    name_counts: dict[str, int] = {}

    for _, row in df.iterrows():
        name = _clean_str(row.get("Exercise"))
        if not name:
            continue

        norm_name = " ".join(name.split())
        name_key = norm_name.casefold()
        name_counts[name_key] = name_counts.get(name_key, 0) + 1

        target_group = _clean_str(row.get("Target Muscle Group"))
        prime_mover = _clean_str(row.get("Prime Mover Muscle"))
        secondary_muscle = _clean_str(row.get("Secondary Muscle"))
        tertiary_muscle = _clean_str(row.get("Tertiary Muscle"))

        primary_equipment = _clean_str(row.get("Primary Equipment"))
        secondary_equipment = _clean_str(row.get("Secondary Equipment"))

        mechanics = _clean_str(row.get("Mechanics"))
        body_region = _clean_str(row.get("Body Region"))
        movement_1 = _clean_str(row.get("Movement Pattern #1"))
        movement_2 = _clean_str(row.get("Movement Pattern #2"))
        movement_3 = _clean_str(row.get("Movement Pattern #3"))

        force_type = _clean_str(row.get("Force Type"))
        posture = _clean_str(row.get("Posture"))
        grip = _clean_str(row.get("Grip"))
        load_pos = _clean_str(row.get("Load Position (Ending)"))
        laterality = _clean_str(row.get("Laterality"))
        primary_class = _clean_str(row.get("Primary Exercise Classification"))

        short_url = _clean_str(row.get("__short_youtube_url"))
        long_url = _clean_str(row.get("__long_youtube_url"))

        difficulty = _map_difficulty(_clean_str(row.get("Difficulty Level")))
        category = _map_category(body_region, prime_mover, target_group)

        muscle_groups = _dedupe_case_insensitive(
            [
                target_group,
                prime_mover,
                secondary_muscle,
                tertiary_muscle,
            ]
        )

        equipment_needed = _dedupe_case_insensitive([primary_equipment, secondary_equipment])

        # Prefer the short demo as main video; keep long explanation as tip.
        video_url = short_url or long_url

        tips: list[str] = []
        for label, val in [
            ("Classification", primary_class),
            ("Mechanics", mechanics),
            ("Movement", movement_1),
            ("Movement", movement_2),
            ("Movement", movement_3),
            ("Posture", posture),
            ("Grip", grip),
            ("Load position", load_pos),
            ("Laterality", laterality),
            ("Force type", force_type),
        ]:
            if val:
                tips.append(f"{label}: {val}")

        if long_url and long_url != video_url:
            tips.append(f"In-depth video: {long_url}")

        # Create a short synthesized description (since the source has no narrative text)
        desc_parts = _dedupe_case_insensitive([
            primary_class,
            movement_1,
            target_group,
            body_region,
        ])
        description = " · ".join(desc_parts[:4])

        # Stable ID based on several fields to reduce collisions
        id_key = "|".join([
            norm_name,
            primary_equipment,
            prime_mover,
            movement_1,
        ])
        exercise_id = _stable_exercise_id(id_key)

        is_compound = "true" if mechanics.casefold() == "compound" else "false"

        out_rows.append(
            {
                "id": exercise_id,
                "name": norm_name,
                "description": description,
                "category": category,
                "muscle_groups": _pg_array_literal(muscle_groups),
                "equipment_needed": _pg_array_literal(equipment_needed),
                "difficulty": difficulty,
                "instructions": _pg_array_literal([]),
                "video_url": video_url,
                "image_url": "",
                "is_compound": is_compound,
                "is_custom": "false",
                "created_by": "",
                "created_at": now_utc,
                "exercisedb_id": "",
                "gif_url": "",
                "tips": _pg_array_literal(tips),
                "last_synced_at": "",
                "sync_status": "pending",
            }
        )

    out_df = pd.DataFrame(out_rows)

    # De-duplicate IDs if collisions occur
    # (rare, but handle defensively)
    if out_df["id"].duplicated().any():
        seen: dict[str, int] = {}
        new_ids: list[str] = []
        for eid in out_df["id"].tolist():
            seen[eid] = seen.get(eid, 0) + 1
            if seen[eid] == 1:
                new_ids.append(eid)
            else:
                new_ids.append(_stable_exercise_id(f"{eid}:{seen[eid]}"))
        out_df["id"] = new_ids

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_csv(OUT_CSV, index=False, quoting=csv.QUOTE_MINIMAL)

    # Audit report
    original_na = (df.drop(columns=["__short_youtube_url", "__long_youtube_url"], errors="ignore").isna().mean() * 100).sort_values(ascending=False)
    out_na = (out_df.replace({"": pd.NA}).isna().mean() * 100).sort_values(ascending=False)

    dup_names = out_df["name"].astype(str).str.strip().str.casefold().duplicated().sum()

    lines: list[str] = []
    lines.append("# Functional Fitness Exercises audit\n")
    lines.append(f"Source: `{XLSX_PATH.relative_to(ROOT)}`")
    lines.append(f"Output (Supabase import): `{OUT_CSV.relative_to(ROOT)}`\n")

    lines.append("## Summary\n")
    lines.append(f"- Source rows read: **{len(df)}**")
    lines.append(f"- Output rows written: **{len(out_df)}**")
    lines.append(f"- Duplicate exercise names (case-insensitive) in output: **{int(dup_names)}**")
    lines.append(f"- YouTube hyperlinks found: short **{link_counts.short_links}**, in-depth **{link_counts.long_links}**\n")

    lines.append("## Source missingness (top 15)\n")
    lines.append("| column | missing % |\n|---|---:|")
    for col, pct in original_na.head(15).items():
        lines.append(f"| {col} | {pct:.1f} |")

    lines.append("\n## Output missingness (top 15)\n")
    lines.append("| column | missing % |\n|---|---:|")
    for col, pct in out_na.head(15).items():
        lines.append(f"| {col} | {pct:.1f} |")

    lines.append("\n## Notes\n")
    lines.append("- `instructions` are empty because the source database does not contain step-by-step text.")
    lines.append("- `video_url` is populated from the *hyperlink targets* behind the YouTube columns (not the displayed 'Video Demonstration' text).")
    lines.append("- `tips` embeds extra metadata (movement patterns, grip, posture, etc.) so you don’t lose richness without changing your table schema.")

    OUT_AUDIT.write_text("\n".join(lines), encoding="utf-8")

    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_AUDIT}")


if __name__ == "__main__":
    main()
